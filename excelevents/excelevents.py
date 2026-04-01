# excelevents.py
import asyncio
import csv
import io
import zipfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Dict, List, Tuple
import aiohttp

import discord
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from redbot.core import commands, Config, data_manager
from redbot.core.bot import Red


class ExcelEvents(commands.Cog):
    """ExcelEvents – Bulk Discord Scheduled Events from Excel (.xlsx) with full image support (2026 polished edition)."""

    MAX_ROWS = 500
    MAX_IMAGE_SIZE = 15 * 1024 * 1024

    def __init__(self, bot: Red):
        self.bot = bot
        self.config = Config.get_conf(self, identifier=9876543210987654321, force_registration=True)
        defaults_guild = {
            "event_mappings": {},
            "last_synced": None,
            "announcement_mode": False,
            "announcement_channel": None,
            "reminder_mode": False,
            "reminder_channel": None,
            "reminder_minutes": [60, 15, 5],
            "reminder_sent": {},
        }
        self.config.register_guild(**defaults_guild)
        self.reminder_task = None
        self.session = None

    async def cog_load(self):
        self.session = aiohttp.ClientSession()
        if self.reminder_task is None or self.reminder_task.done():
            self.reminder_task = asyncio.create_task(self._reminder_task())

    def cog_unload(self):
        if self.session:
            asyncio.create_task(self.session.close())
        if self.reminder_task and not self.reminder_task.done():
            self.reminder_task.cancel()

    # ====================== REFINED IMAGE DOWNLOADER ======================
    async def _download_image(self, url: str) -> Optional[bytes]:
        if not url or not str(url).startswith(("http://", "https://")):
            return None

        url = url.strip()

        if "imgur.com" in url:
            url = url.replace(".jpeg", ".jpg").replace(".JPEG", ".jpg")
            if "i.imgur.com" not in url and "imgur.com" in url:
                image_id = url.split("/")[-1].split("?")[0].split(".")[0]
                url = f"https://i.imgur.com/{image_id}.jpg"
            if "?" in url:
                url = url.split("?")[0]

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
            "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
            "Referer": "https://imgur.com/",
        }

        timeout = aiohttp.ClientTimeout(total=25)

        for attempt in range(2):
            try:
                async with self.session.get(url, headers=headers, timeout=timeout, allow_redirects=True) as resp:
                    if resp.status != 200:
                        continue

                    content_length_header = resp.headers.get("Content-Length")
                    if content_length_header is not None:
                        try:
                            content_length = int(content_length_header)
                            if content_length > self.MAX_IMAGE_SIZE:
                                continue
                        except ValueError:
                            pass

                    data = await resp.read()

                    if len(data) > self.MAX_IMAGE_SIZE:
                        continue

                    content_type = resp.headers.get("Content-Type", "").lower()
                    if len(data) > 1024 and (
                        any(x in content_type for x in ("image/", "jpeg", "png", "gif", "webp")) or
                        data.startswith((b'\xff\xd8', b'\x89PNG', b'GIF8', b'RIFF'))
                    ):
                        return data
            except Exception:
                if attempt == 0:
                    await asyncio.sleep(1.5)
                continue
        return None

    async def _parse_datetime(self, value) -> Optional[datetime]:
        if not value:
            return None
        if isinstance(value, (int, float)):
            try:
                base = datetime(1899, 12, 30)
                dt = base + timedelta(days=value)
                return dt.replace(tzinfo=timezone.utc)
            except Exception:
                pass
        value_str = str(value).strip()
        if not value_str:
            return None
        formats = [
            "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S",
            "%m/%d/%y %H:%M", "%m/%d/%y %H:%M:%S",
            "%m/%d/%Y %I:%M %p", "%m/%d/%Y %I:%M:%S %p",
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %I:%M %p",
            "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
        ]
        for fmt in formats:
            try:
                dt = datetime.strptime(value_str, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
        return None

    def _normalize_key(self, name: str) -> str:
        return str(name).strip().lower()

    def _is_valid_xlsx(self, file_path: Path) -> bool:
        try:
            with open(file_path, "rb") as f:
                header = f.read(4)
            return header[:2] == b'PK'
        except Exception:
            return False

    def _get_column_indices(self, headers: List[str]) -> Dict[str, int]:
        col_map = {}
        aliases = {
            "name": ["name", "event name", "title", "event"],
            "start": ["start", "start time", "start date", "date", "when"],
            "end": ["end", "end time", "end date"],
            "description": ["description", "desc", "details"],
            "type": ["type", "event type", "format", "kind"],
            "location": ["location", "place", "venue", "address", "link"],
            "channelid": ["channelid", "channel id", "channel", "voice channel", "stage channel"],
            "image": ["image", "cover", "banner", "imageurl", "cover image", "event image"],
        }
        for i, h in enumerate(headers):
            if not h:
                continue
            norm = self._normalize_key(h)
            for canonical, alias_list in aliases.items():
                if any(norm == self._normalize_key(a) for a in alias_list):
                    col_map[canonical] = i
                    break
            else:
                col_map[norm] = i
        return col_map

    def _get_cell(self, row: tuple, col_map: Dict[str, int], key: str, default=None):
        idx = col_map.get(key)
        if idx is not None and idx < len(row):
            val = row[idx]
            return val if val is not None else default
        return default

    async def _create_event_with_image(self, guild: discord.Guild, data: Dict, image_bytes: Optional[bytes] = None) -> Optional[discord.ScheduledEvent]:
        name = str(data.get("name", "")).strip()
        if not name or len(name) > 100:
            return None

        start_time = await self._parse_datetime(data.get("start"))
        if not start_time:
            return None

        end_time = await self._parse_datetime(data.get("end"))
        description = str(data.get("description", "")).strip()[:1000] or None
        event_type_str = str(data.get("type", "")).strip().lower() or "voice"
        location = str(data.get("location", "")).strip() or None
        channel_id_input = data.get("channelid")

        if event_type_str in ("external", "url", "link"):
            entity_type = discord.EntityType.external
        elif event_type_str == "stage":
            entity_type = discord.EntityType.stage_instance
        else:
            entity_type = discord.EntityType.voice

        channel = None
        if entity_type in (discord.EntityType.voice, discord.EntityType.stage_instance) and channel_id_input:
            try:
                ch_id = int(str(channel_id_input).strip())
                temp_ch = guild.get_channel(ch_id)
                if temp_ch and (
                    (entity_type == discord.EntityType.voice and isinstance(temp_ch, discord.VoiceChannel)) or
                    (entity_type == discord.EntityType.stage_instance and isinstance(temp_ch, discord.StageChannel))
                ):
                    channel = temp_ch
            except Exception:
                pass

        create_kwargs = {
            "name": name,
            "description": description,
            "start_time": start_time,
            "end_time": end_time,
            "entity_type": entity_type,
            "privacy_level": discord.PrivacyLevel.guild_only,
        }

        if image_bytes:
            create_kwargs["image"] = image_bytes

        try:
            if entity_type == discord.EntityType.external:
                if not location:
                    return None
                create_kwargs["location"] = location
            else:
                if not channel:
                    return None
                create_kwargs["channel"] = channel

            event = await guild.create_scheduled_event(**create_kwargs)
            await asyncio.sleep(2.0)

            if image_bytes:
                try:
                    event = await guild.fetch_scheduled_event(event.id)
                    if not event.cover_image:
                        await event.edit(image=image_bytes)
                        await asyncio.sleep(1.5)
                except Exception:
                    pass

            await asyncio.sleep(1.0)
            return event
        except Exception:
            return None

    async def _update_event(self, event: discord.ScheduledEvent, data: Dict, image_bytes: Optional[bytes] = None):
        try:
            edit_kwargs = {
                "name": str(data.get("name", "")).strip(),
                "description": str(data.get("description", "")).strip()[:1000] or None,
                "start_time": await self._parse_datetime(data.get("start")),
                "end_time": await self._parse_datetime(data.get("end")),
            }
            if image_bytes:
                edit_kwargs["image"] = image_bytes
            await event.edit(**edit_kwargs)
            await asyncio.sleep(1.2)
        except Exception:
            pass

    def _create_event_embed(self, event: discord.ScheduledEvent) -> discord.Embed:
        embed = discord.Embed(
            title=event.name[:256],
            description=(event.description or "No description provided.")[:4096],
            color=discord.Color.blurple(),
            url=event.url,
        )
        if event.start_time:
            ts = int(event.start_time.timestamp())
            embed.add_field(name="Start", value=f"<t:{ts}:F> (<t:{ts}:R>)", inline=True)
        if event.end_time:
            ts = int(event.end_time.timestamp())
            embed.add_field(name="End", value=f"<t:{ts}:F>", inline=True)

        loc = event.location or (event.channel.mention if event.channel else "Voice/Stage")
        embed.add_field(name="Location", value=loc, inline=False)
        embed.add_field(name="Type", value=event.entity_type.name.replace("_", " ").title(), inline=True)
        embed.set_footer(text="New Event • Synced via ExcelEvents • RedBot 2026")
        return embed

    def _create_reminder_embed(self, event: discord.ScheduledEvent, minutes: int) -> discord.Embed:
        embed = discord.Embed(
            title=f"{event.name} starts in {minutes} minutes!",
            description=(event.description or "")[:4096],
            color=discord.Color.orange(),
            url=event.url,
        )
        if event.start_time:
            ts = int(event.start_time.timestamp())
            embed.add_field(name="Exact Time", value=f"<t:{ts}:F>", inline=False)
        loc = event.location or (event.channel.mention if event.channel else "Voice/Stage")
        embed.add_field(name="Location", value=loc, inline=False)
        embed.set_footer(text=f"Reminder • ExcelEvents")
        return embed

    async def _validate_excel(self, file_path: Path, guild: discord.Guild) -> Tuple[List[str], List[str]]:
        errors: List[str] = []
        warnings: List[str] = []

        if not file_path.exists():
            errors.append("No events.xlsx file found. Use `upload` or `paste` first.")
            return errors, warnings

        if file_path.stat().st_size == 0:
            errors.append("The uploaded file is empty.")
            return errors, warnings

        is_real_xlsx = self._is_valid_xlsx(file_path)

        try:
            if not is_real_xlsx:
                raise zipfile.BadZipFile("Not a valid .xlsx")

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            if ws is None or ws.max_row < 1:
                errors.append("Worksheet is empty or unreadable.")
                return errors, warnings

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
            col_map = self._get_column_indices(headers)

        except (zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException):
            errors.append("This is **not** a valid .xlsx file. Use `paste` instead.")
            return errors, warnings
        except Exception as e:
            errors.append(f"Failed to read file: {type(e).__name__} – {e}")
            return errors, warnings

        required = {"name", "start"}
        missing = [col for col in required if col not in col_map]
        if missing:
            errors.append(f"Missing required column(s): {', '.join(missing)}")

        row_num = 1
        seen_names: set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or all(v is None for v in row):
                continue

            name = str(self._get_cell(row, col_map, "name", "")).strip()
            if not name:
                errors.append(f"Row {row_num}: Missing or empty **Name**")
                continue

            if len(name) > 100:
                errors.append(f"Row {row_num}: Name too long (max 100 characters)")

            key = self._normalize_key(name)
            if key in seen_names:
                warnings.append(f"Row {row_num}: Duplicate name '{name}'")
            seen_names.add(key)

            start_dt = await self._parse_datetime(self._get_cell(row, col_map, "start"))
            if not start_dt:
                errors.append(f"Row {row_num}: Invalid **Start** time format")
            elif start_dt < datetime.now(timezone.utc):
                warnings.append(f"Row {row_num}: Start time is in the past")

        if not seen_names:
            errors.append("No valid event rows found in the file.")

        return errors, warnings

    async def _reminder_task(self):
        await self.bot.wait_until_ready()
        while True:
            try:
                for guild in self.bot.guilds:
                    config = self.config.guild(guild)
                    await asyncio.sleep(60)
            except Exception:
                await asyncio.sleep(60)

    # ====================== MAIN COMMAND GROUP ======================
    @commands.group(invoke_without_command=True)
    async def excelevents(self, ctx: commands.Context):
        """Bulk manage Discord Scheduled Events using Excel files.
        
        Use ,excelevents template to get the Excel template.
        """
        if ctx.invoked_subcommand is None:
            await ctx.send_help(ctx.command)

    # ====================== PROFESSIONAL EXCEL TEMPLATE ======================
    @excelevents.command(name="template")
    async def template(self, ctx: commands.Context):
        """Uploads a ready-to-use Excel (.xlsx) template suitable for excelevents."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"

        headers = ["name", "start", "end", "description", "type", "location", "channelid", "image"]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = 25

        example = [
            "Example Movie Night",
            "2026-04-05 20:00",
            "2026-04-05 22:00",
            "Join us for a movie night in voice chat! Popcorn not included 🍿",
            "voice",
            "",
            "123456789012345678",
            "https://i.imgur.com/3eQczTs.jpg"
        ]

        for col, value in enumerate(example, 1):
            ws.cell(row=2, column=col, value=value)

        ws2 = wb.create_sheet("README")
        ws2['A1'] = "Excelevents Excel Template – How to use"
        ws2['A1'].font = Font(bold=True, size=14)

        instructions = [
            ("1.", "Fill one event per row starting from row 2"),
            ("2.", "Required: name + start time"),
            ("3.", "Type → voice | stage | external"),
            ("4.", "External events → put the link in the 'location' column"),
            ("5.", "Voice/Stage events → put channel ID in the 'channelid' column"),
            ("6.", "Image column → direct image link (Imgur works best)"),
            ("7.", "Save the file → then run ,excelevents upload and attach it"),
        ]

        for i, (num, text) in enumerate(instructions, start=3):
            ws2[f'A{i}'] = num
            ws2[f'B{i}'] = text

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file = discord.File(buffer, filename="excelevents_template.xlsx")

        embed = discord.Embed(
            title="📊 ExcelEvents Template",
            description="**Download the attached `.xlsx` file!**\n\nFill it out → then use `,excelevents upload`",
            color=discord.Color.green()
        )

        await ctx.send(embed=embed, file=file)

    @excelevents.command(name="upload")
    async def upload(self, ctx: commands.Context):
        """Upload an .xlsx file to be used for events."""
        if not ctx.message.attachments:
            await ctx.send("Please attach an `.xlsx` or `.xls` file.")
            return
        attachment = ctx.message.attachments[0]
        if not attachment.filename.lower().endswith((".xlsx", ".xls")):
            await ctx.send("Only `.xlsx` or `.xls` files are supported.")
            return

        data_path: Path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        if file_path.exists():
            file_path.unlink()

        await attachment.save(str(file_path))
        await ctx.send("File uploaded (old file replaced). Use `check`.")

    @excelevents.command(name="paste")
    async def paste(self, ctx: commands.Context):
        """Paste CSV data to create the events file."""
        lines = ctx.message.content.splitlines()
        csv_text = "\n".join(lines[1:]) if len(lines) > 1 else ""

        if not csv_text.strip():
            await ctx.send("Please paste CSV data after the command.")
            return

        data_path = data_manager.cog_data_path(self)
        data_path.mkdir(parents=True, exist_ok=True)
        file_path = data_path / "events.xlsx"

        if file_path.exists():
            file_path.unlink()

        try:
            input_io = io.StringIO(csv_text.strip())
            reader = csv.reader(input_io, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL, skipinitialspace=True)
            rows = [[cell.strip() for cell in row] for row in reader if row and any(cell.strip() for cell in row)]

            if len(rows) < 1:
                await ctx.send("No valid rows found.")
                return
            if len(rows) - 1 > self.MAX_ROWS:
                rows = rows[:self.MAX_ROWS + 1]
                await ctx.send(f"Only first {self.MAX_ROWS} events saved.")

            if rows:
                header_len = len(rows[0])
                for i in range(1, len(rows)):
                    rows[i] += [''] * (header_len - len(rows[i]))

            wb = openpyxl.Workbook()
            ws = wb.active
            for row in rows:
                ws.append(row)
            wb.save(file_path)

            await ctx.send(f"CSV saved! **{len(rows)-1}** events loaded.\nUse `check`.")
        except Exception as e:
            await ctx.send(f"Failed to parse CSV: {type(e).__name__} – {e}")

    @excelevents.command(name="check")
    async def check(self, ctx: commands.Context):
        """Validate the events file before syncing."""
        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        await ctx.send("Running validation...")
        errors, warnings = await self._validate_excel(file_path, ctx.guild)

        if errors:
            await ctx.send("**Validation Failed:**\n" + "\n".join(f"Error: {msg}" for msg in errors))
        elif warnings:
            await ctx.send("**Valid with warnings:**\n" + "\n".join(f"Warning: {msg}" for msg in warnings) + "\n\nYou may now run `sync`.")
        else:
            await ctx.send("**Perfect!** Ready to sync.")

    @excelevents.command(name="sync")
    async def sync(self, ctx: commands.Context):
        """Sync the spreadsheet to Discord Scheduled Events (create/update/delete + images)."""
        if not ctx.guild.me.guild_permissions.manage_events:
            await ctx.send("I need the **Manage Events** permission.")
            return

        data_path = data_manager.cog_data_path(self)
        file_path = data_path / "events.xlsx"
        if not file_path.exists():
            await ctx.send("No file found. Use `upload` or `paste` first.")
            return

        errors, warnings = await self._validate_excel(file_path, ctx.guild)
        if errors:
            await ctx.send("Validation failed. Run `check` first.")
            return

        await ctx.send("Syncing events with image support...")

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
            col_map = self._get_column_indices(headers)

            global_image_bytes = None
            if ctx.message.attachments:
                att = ctx.message.attachments[0]
                if att.content_type and att.content_type.startswith("image/") and att.size < self.MAX_IMAGE_SIZE:
                    global_image_bytes = await att.read()

            mappings = await self.config.guild(ctx.guild).event_mappings()
            new_mappings: Dict[str, int] = {}
            active_keys = set()
            processed = 0
            new_events_created = []

            row_num = 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or all(v is None for v in row):
                    continue

                name = str(self._get_cell(row, col_map, "name", "")).strip()
                if not name:
                    continue

                key = self._normalize_key(name)
                active_keys.add(key)

                data = {
                    "name": name,
                    "start": self._get_cell(row, col_map, "start"),
                    "end": self._get_cell(row, col_map, "end"),
                    "description": self._get_cell(row, col_map, "description"),
                    "type": self._get_cell(row, col_map, "type"),
                    "location": self._get_cell(row, col_map, "location"),
                    "channelid": self._get_cell(row, col_map, "channelid"),
                }

                image_url = str(self._get_cell(row, col_map, "image", "")).strip()
                image_bytes = None

                if image_url:
                    image_bytes = await self._download_image(image_url)
                    if image_bytes:
                        await ctx.send(f"Row {row_num}: Image downloaded ({len(image_bytes)//1024} KB) for **{name}**")
                    else:
                        await ctx.send(f"Row {row_num}: Image failed for **{name}** — event created without cover")
                elif global_image_bytes:
                    image_bytes = global_image_bytes
                    await ctx.send(f"Row {row_num}: Using attached image for **{name}**")

                start_time = await self._parse_datetime(data["start"])
                if not start_time:
                    continue

                if key in mappings:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(mappings[key])
                        await self._update_event(event, data, image_bytes)
                        new_mappings[key] = event.id
                        processed += 1
                        continue
                    except Exception:
                        pass

                new_event = await self._create_event_with_image(ctx.guild, data, image_bytes)
                if new_event:
                    new_mappings[key] = new_event.id
                    new_events_created.append(new_event)
                    processed += 1
                else:
                    await ctx.send(f"Failed to create event: {name}")

            deleted = 0
            for old_key, old_id in list(mappings.items()):
                if old_key not in active_keys:
                    try:
                        event = await ctx.guild.fetch_scheduled_event(old_id)
                        await event.delete()
                        deleted += 1
                    except Exception:
                        pass

            await self.config.guild(ctx.guild).event_mappings.set(new_mappings)
            await self.config.guild(ctx.guild).last_synced.set(datetime.now(timezone.utc).isoformat())

            summary = f"✅ **Sync complete!**\n• Processed: {processed}\n• Created: {len(new_events_created)}\n• Deleted: {deleted}"
            await ctx.send(summary)

            if new_events_created:
                for event in new_events_created[:5]:
                    await ctx.send(embed=self._create_event_embed(event))

        except Exception as e:
            await ctx.send(f"Sync failed: {type(e).__name__} – {e}")


async def setup(bot):
    await bot.add_cog(ExcelEvents(bot))
